import os, re
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook

# ===== 你指定的配置（只保留这四列） =====
DATE_RE         = re.compile(r"\b\d{4}-\d{2}-\d{2}\b")     # YYYY-MM-DD
VENDOR_PART_RE  = re.compile(r"\b(?=[A-Z0-9-]*[A-Z])(?=[A-Z0-9-]*-)[A-Z0-9-]+\b")
PO_OK_RE        = re.compile(r"[A-Za-z0-9-]+")             # 允许 PO=0 / 字母数字 / 短横线
HEADERS         = ["_file", "RTV DATE", "VENDOR PART #", "PO #", "QTY"]
SHEET_NAME      = "RTV Data"

# 仅用于定位价格行（不导出 cost）
MONEY_RE        = re.compile(r"\$\s?\d{1,3}(?:,\d{3})*\.\d{2}")

def norm(s: str) -> str:
    # 兼容 "$ 1,234.56"：把 "$ " 合并成 "$"
    s = (s or "").replace("$ ", "$")
    return re.sub(r"\s+", " ", s.strip())

def find_rtv_date(lines):
    for line in lines:
        if "RTV DATE" in line.upper():
            m = DATE_RE.search(line)
            if m: return m.group(0)
    for line in lines[:80]:
        m = DATE_RE.search(line)
        if m: return m.group(0)
    return ""

def find_table_bounds(lines):
    """表头可跨两行：把相邻两行拼接来找起点；表尾到 MERCHANDISE TOTAL / RTV GRAND TOTAL 为止。"""
    start = end = None
    n = len(lines)
    for i in range(n):
        a = norm(lines[i]).upper()
        b = norm(lines[i+1]).upper() if i+1 < n else ""
        joined = a + " " + b
        if start is None and ("VENDOR" in joined and "PART #" in joined and "PO #" in joined and "QTY" in joined):
            # 如果当前行已经含 PO#/QTY，明细从下一行开始；否则从再下一行开始
            start = i + 1 if ("PO #" in a or "QTY" in a) else (i + 2)
            continue
        if start is not None:
            if "MERCHANDISE TOTAL" in a or "RTV GRAND TOTAL" in a:
                end = i
                break
    return start, end

def find_vendor_in_text(text: str) -> str:
    """从一行文本里找第一个像 'FFSGS6260-36' 这种 Vendor Part。"""
    parts = text.split()
    for t in parts:
        if VENDOR_PART_RE.fullmatch(t):
            return t
    # 兜底：有些会把描述粘在后面，例如 FFSGS6274-30GIOVANNI
    m = re.search(r"[A-Z0-9]{3,}(?:-[A-Z0-9]{2,})+", text)
    return m.group(0) if m else ""

def parse_price_line_get_po_qty(parts):
    """
    用金额作为锚点定位 PO/QTY：
      ... PO  QTY  UOM  UNIT_COST  EXT_COST
    """
    money_idx = [i for i, p in enumerate(parts) if MONEY_RE.match(p)]
    if len(money_idx) < 2:
        return None
    pos_unit, pos_ext = money_idx[-2], money_idx[-1]
    if pos_unit < 3:
        return None
    po  = parts[pos_unit - 3]
    qty = parts[pos_unit - 2]
    if not (PO_OK_RE.fullmatch(po) and re.fullmatch(r"\d+", qty)):
        return None
    return po, qty

def extract_rows_from_page(lines, src_file):
    rows = []
    rtv_date = find_rtv_date(lines)
    start, end = find_table_bounds(lines)
    if start is None or end is None:
        return rows

    last_vendor = ""  # 记住最近一次出现的 Vendor，供下一行价格行配对
    i = start
    while i < end:
        s = norm(lines[i])
        if not s:
            i += 1; continue
        parts = s.split()

        # 1) 先看看这一行有没有 Vendor
        v = find_vendor_in_text(s)
        if v:
            last_vendor = v

        # 2) 再看看这一行是不是“价格行”（含两笔金额），如果是，用金额锚点取 PO/QTY
        pq = parse_price_line_get_po_qty(parts)
        if pq and last_vendor:
            po, qty = pq
            rows.append([src_file, rtv_date, last_vendor, po, qty])
            last_vendor = ""     # 成功配对后清空，避免串行
            i += 1
            continue

        # 3) 否则继续往下扫
        i += 1

    return rows

# ===== 交互：选择 PDF 文件夹 + 选择/新建 Excel（结果会追加到此文件） =====
root = tk.Tk(); root.withdraw()
pdf_folder = filedialog.askdirectory(title="选择包含PDF文件的文件夹")
if not pdf_folder:
    messagebox.showwarning("未选择文件夹", "程序退出。"); raise SystemExit

xlsx_path = filedialog.asksaveasfilename(
    title="选择/新建 Excel（结果会追加到该文件）",
    defaultextension=".xlsx",
    filetypes=[("Excel 文件", "*.xlsx")]
)
if not xlsx_path:
    messagebox.showwarning("未选择保存路径", "程序退出。"); raise SystemExit

# ===== 扫描并提取 =====
all_rows = []
for fname in sorted(os.listdir(pdf_folder)):
    if not fname.lower().endswith(".pdf"): continue
    fpath = os.path.join(pdf_folder, fname)
    try:
        with pdfplumber.open(fpath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                lines = text.splitlines()
                all_rows.extend(extract_rows_from_page(lines, fname))
    except Exception as e:
        print(f"[WARN] 读取失败 {fname}: {e}")

if not all_rows:
    messagebox.showinfo("结果", "未提取到任何记录。"); raise SystemExit

# ===== 追加/新建 Excel =====
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    if ws.max_row == 1 and (ws.cell(1,1).value or "") != HEADERS[0]:
        ws.append(HEADERS)
else:
    wb = Workbook()
    default = wb.active; wb.remove(default)
    ws = wb.create_sheet(SHEET_NAME)
    ws.append(HEADERS)

for r in all_rows:
    ws.append(r)

wb.save(xlsx_path)
messagebox.showinfo("完成", f"✅ 本次追加 {len(all_rows)} 条记录\n已保存到：{xlsx_path}")
