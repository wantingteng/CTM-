import pandas as pd
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# === UI 文件选择 ===
root = tk.Tk()
root.withdraw()

# 选择主报表 Excel 文件
master_file = filedialog.askopenfilename(
    title="请选择主报表 Excel 文件",
    filetypes=[("Excel Files", "*.xlsx")]
)
if not master_file:
    messagebox.showwarning("未选择主表", "程序终止。")
    exit()

# 选择要追加的新 CSV 数据文件
new_data_csv = filedialog.askopenfilename(
    title="请选择要追加的数据文件（CSV）",
    filetypes=[("CSV Files", "*.csv")]
)
if not new_data_csv:
    messagebox.showwarning("未选择CSV文件", "程序终止。")
    exit()

# === 加载数据 ===
df_new = pd.read_csv(new_data_csv, dtype=str)
df_new.drop(columns=["PACKAGE DATE", "PACKAGE NUMBER", "PO", "PART"], errors="ignore", inplace=True)

# 统一 RTV PO DATE 格式：转为 5/5/2025 这种格式
df_new["RTV PO DATE"] = pd.to_datetime(df_new["RTV PO DATE"], errors="coerce")
df_new.insert(1, "MONTH", df_new["RTV PO DATE"].dt.month.astype("Int64"))  # 插入月份列
df_new["RTV PO DATE"] = df_new["RTV PO DATE"].apply(lambda d: f"{d.month}/{d.day}/{d.year}" if pd.notna(d) else "")

# 添加 STATUS 列（如果不存在）
if "STATUS" not in df_new.columns:
    df_new["STATUS"] = ""

# === 加载主表 ===
if os.path.exists(master_file):
    df_master = pd.read_excel(master_file, dtype=str, engine="openpyxl")
    for col in df_new.columns:
        if col not in df_master.columns:
            df_master[col] = ""  # 保持格式统一
    df_combined = pd.concat([df_master, df_new[df_master.columns]], ignore_index=True)
else:
    df_combined = df_new

# === 日期排序（不会丢失自定义列） ===
if "RTV PO DATE" in df_combined.columns:
    df_combined["_RTV_DATE_OBJ"] = pd.to_datetime(df_combined["RTV PO DATE"], errors="coerce")
    df_combined = df_combined.sort_values("_RTV_DATE_OBJ").drop(columns=["_RTV_DATE_OBJ"]).reset_index(drop=True)

# === 写入 Excel 并设置货币格式 ===
df_combined.to_excel(master_file, index=False, engine="openpyxl")

wb = load_workbook(master_file)
ws = wb.active

# 自动识别金额列并设置货币格式
money_columns = [i + 1 for i, cell in enumerate(ws[1]) if cell.value and any(k in str(cell.value).upper() for k in ["TOTAL", "AMOUNT", "COST"])]

for row in ws.iter_rows(min_row=2):
    for col_idx in money_columns:
        cell = row[col_idx - 1]
        try:
            val = float(str(cell.value).replace(",", "").replace("$", ""))
            cell.value = val
            cell.number_format = '"$"#,##0.00'
        except:
            continue

wb.save(master_file)

messagebox.showinfo("完成", f"✅ 数据已成功追加并保存至主表：\n{master_file}")
