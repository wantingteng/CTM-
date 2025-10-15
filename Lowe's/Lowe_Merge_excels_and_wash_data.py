import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import subprocess

# 1. 三种模式可选（图形界面选择）：
# 文件夹模式（批量扫描）：自动扫描选中的整个文件夹下的所有 Excel 文件。
#
# 文件模式（手动多选文件）：弹窗选择多个具体的 Excel 文件合并。
#
# 追加模式（向已有 merged 文件追加数据）：选择一个已有的 merged.xlsx 文件，然后选新文件追加数据进去。



EXCEL_EXTENSIONS = ['.xlsx', '.xls', '.xlsm', '.xlsb']
COLUMNS_TO_DROP = [
    "Contact Email", "Chargeback Comments", "Item Description", "Line",
    "Tracking No", "Selling Location", "Location", "SCAC", "PRO#",
    "Authorization ID", "Rtm Vendor Name", "Item #", "Rtm Vendor #","Contact Name"
]
CURRENCY_COLUMNS = [
    "Deduction Amount", "RTM Total Cost", "Unit Cost",
    "Extended Cost", "Chargeback Amount", "approuved amount"
]
STANDARD_COLUMNS = [
    "Debit Memo #", "Debit Memo Date", "Deduction Amount", "RTM Date", "RTM Total Cost",
    "SOS Po", "RTV Reason", "RTV method", "Lowes Quantity", "Unit Cost",
    "Extended Cost", "Chargeback Amount", "Chargeback Reason", "approuved amount"
]

def get_all_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if any(file.lower().endswith(ext) for ext in EXCEL_EXTENSIONS):
                excel_files.append(os.path.join(root, file))
    return excel_files

def read_and_clean_excel(file_path):
    try:
        df = pd.read_excel(file_path)
        cols = [c.lower().strip() for c in df.columns]
        if 'deduction number' in cols and 'deduction date' in cols and 'deduction amount' in cols:
            df.columns = cols
            result = pd.DataFrame()
            result["Debit Memo #"] = df["deduction number"]
            result["Debit Memo Date"] = pd.to_datetime(df["deduction date"], errors='coerce').dt.date
            result["Deduction Amount"] = df["deduction amount"]
            for col in STANDARD_COLUMNS:
                if col not in result.columns:
                    if col == 'approuved amount':
                        result[col] = ''
                    elif col in CURRENCY_COLUMNS:
                        result[col] = 0
                    else:
                        result[col] = ''
            return result[STANDARD_COLUMNS]
        df = df.drop(columns=[col for col in COLUMNS_TO_DROP if col in df.columns], errors='ignore')
        return df
    except Exception as e:
        print(f"读取失败: {file_path}, 错误: {e}")
        return None

def apply_currency_format(file_path, currency_columns):
    wb = load_workbook(file_path)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for col_name in currency_columns:
        if col_name in header:
            col_idx = header.index(col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if col_name == "approuved amount":
                        cell.number_format = '"$"#,##0.00'
                    elif isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0.00'
    wb.save(file_path)
    wb.close()

def reorder_columns(df):
    cols = list(df.columns)
    # 确保 Month 在 Date 后
    if 'Debit Memo Date' in cols and 'Debit Memo Month' in cols:
        date_idx = cols.index('Debit Memo Date')
        cols.remove('Debit Memo Month')
        cols.insert(date_idx + 1, 'Debit Memo Month')

    # ⭐ 新增逻辑：确保 RTM # 在 RTM Date 后
    if 'RTM Date' in cols and 'RTM #' in cols:
        date_idx = cols.index('RTM Date')
        cols.remove('RTM #')
        cols.insert(date_idx + 1, 'RTM #')

    df = df[cols]
    return df


def process_new_data(df):
    if 'Debit Memo Date' in df.columns:
        df['Debit Memo Date'] = pd.to_datetime(df['Debit Memo Date'], errors='coerce')
        df = df.sort_values(by='Debit Memo Date', kind='stable', ascending=True)
        df['Debit Memo Month'] = df['Debit Memo Date'].dt.strftime('%m')
        df['Debit Memo Date'] = df['Debit Memo Date'].dt.date
    deduction_mask = df['Debit Memo #'].astype(str).str.startswith("DMRL", na=False)
    deduction_data = df[deduction_mask]
    normal_data = df[~deduction_mask]
    deduction_data = deduction_data.drop_duplicates(subset=['Debit Memo #', 'Debit Memo Date', 'Deduction Amount'])
    df = pd.concat([normal_data, deduction_data], ignore_index=True)
    return df

def open_folder(path):
    folder = os.path.dirname(path)
    if os.name == 'nt':
        subprocess.Popen(f'explorer "{folder}"')

def merge_files(file_list):
    merged_df = pd.DataFrame()
    for file in file_list:
        df = read_and_clean_excel(file)
        if df is not None:
            if 'approuved amount' not in df.columns:
                df['approuved amount'] = ''
            df = process_new_data(df)
            merged_df = pd.concat([merged_df, df], ignore_index=True)
    merged_df = reorder_columns(merged_df)
    merged_df['Debit Memo Date'] = pd.to_datetime(merged_df['Debit Memo Date'], errors='coerce')
    merged_df = merged_df.sort_values(by='Debit Memo Date', kind='stable', ascending=True)
    merged_df['Debit Memo Date'] = merged_df['Debit Memo Date'].dt.date

    if merged_df.empty:
        messagebox.showwarning("提示", "没有成功读取任何数据。")
        return
    output_file = filedialog.asksaveasfilename(
        title="保存合并后的Excel文件",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if output_file:
        try:
            merged_df.to_excel(output_file, index=False)
            apply_currency_format(output_file, CURRENCY_COLUMNS)
            open_folder(output_file)
            messagebox.showinfo("完成", f"合并完成，文件已保存至：\n{output_file}")
        except PermissionError:
            messagebox.showerror("错误", f"无法保存文件：{output_file}\n请关闭该文件后重试。")

def folder_mode():
    folder_selected = filedialog.askdirectory(title="请选择包含Excel文件的文件夹")
    if folder_selected:
        files = get_all_excel_files(folder_selected)
        if files:
            merge_files(files)
        else:
            messagebox.showwarning("提示", "文件夹中未找到Excel文件。")

def file_mode():
    files = filedialog.askopenfilenames(
        title="请选择要合并的Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.xlsb")]
    )
    if files:
        merge_files(files)

def append_mode():
    merged_file = filedialog.askopenfilename(
        title="请选择已有的merged.xlsx文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not merged_file:
        return
    existing_df = pd.read_excel(merged_file)
    if 'Debit Memo Date' in existing_df.columns:
        existing_df['Debit Memo Date'] = existing_df['Debit Memo Date'].astype(str)
        if 'Debit Memo Month' not in existing_df.columns:
            existing_df['Debit Memo Month'] = pd.to_datetime(existing_df['Debit Memo Date'], errors='coerce').dt.strftime('%m')
        existing_df = reorder_columns(existing_df)
    files_to_append = filedialog.askopenfilenames(
        title="请选择要追加的Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.xlsb")]
    )
    if files_to_append:
        appended_dfs = []
        for file in files_to_append:
            df = read_and_clean_excel(file)
            if df is not None:
                if 'approuved amount' not in df.columns:
                    df['approuved amount'] = ''
                df = process_new_data(df)
                appended_dfs.append(df)
        if appended_dfs:
            new_data = pd.concat(appended_dfs, ignore_index=True)
            merged_df = pd.concat([existing_df, new_data], ignore_index=True)
            merged_df = reorder_columns(merged_df)
            merged_df['Debit Memo Date'] = pd.to_datetime(merged_df['Debit Memo Date'], errors='coerce')
            merged_df = merged_df.sort_values(by='Debit Memo Date', kind='stable', ascending=True)
            merged_df['Debit Memo Date'] = merged_df['Debit Memo Date'].dt.date

            try:
                merged_df.to_excel(merged_file, index=False)
                apply_currency_format(merged_file, CURRENCY_COLUMNS)
                open_folder(merged_file)
                messagebox.showinfo("完成", f"追加完成，merged文件已更新：\n{merged_file}")
            except PermissionError:
                messagebox.showerror("错误", f"无法保存文件：{merged_file}\n请关闭该文件后重试。")

def main():
    root = tk.Tk()
    root.title("Lowe's Excel文件合并 & 追加工具（终极v5业务版）")
    tk.Label(root, text="请选择操作模式：", font=("Arial", 14)).pack(pady=10)
    tk.Button(root, text="1️⃣ 文件夹模式（批量扫描）", command=folder_mode, width=40, height=2).pack(pady=5)
    tk.Button(root, text="2️⃣ 文件模式（多选Excel）", command=file_mode, width=40, height=2).pack(pady=5)
    tk.Button(root, text="3️⃣ 追加模式（向已有merged追加）", command=append_mode, width=40, height=2).pack(pady=5)
    root.mainloop()

if __name__ == "__main__":
    main()
