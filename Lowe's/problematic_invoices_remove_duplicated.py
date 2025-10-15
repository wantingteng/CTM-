import pandas as pd
from tkinter import filedialog, Tk
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

#你手动添加的列（如 Status）会被保留；

#新文件没有 Month 也没关系，会重新生成；

#日期自动转换格式，非标准格式也能尽量处理；

#如果有重复的发票号（Invoice Number），只保留一份；

#输出为新的 Excel 文件。
# 设置输出路径
output_dir = r"C:\Reporting\Lowe's\Problematic Invoices\2025"
output_file = os.path.join(output_dir, "combined_cleaned.xlsx")

# 选择文件
def select_file(title="选择文件"):
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

print("请选择母表 Excel 文件：")
main_file = select_file("选择母表文件")
print("请选择需要追加的新表 Excel 文件：")
append_file = select_file("选择追加的新文件")

# 读取文件
df_main = pd.read_excel(main_file)
df_append = pd.read_excel(append_file)

# 确保新表有母表的所有列（如缺则填空），并按母表列顺序排列
for col in df_main.columns:
    if col not in df_append.columns:
        df_append[col] = None
df_append = df_append[df_main.columns]

# 将日期列转换为 datetime 类型
for df in [df_main, df_append]:
    if "Invoice Date" in df.columns:
        df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors='coerce')

# 删除已有 Month 列，重新生成
for df in [df_main, df_append]:
    if "Month" in df.columns:
        df.drop(columns=["Month"], inplace=True)
    if "Invoice Date" in df.columns:
        month_series = df["Invoice Date"].dt.strftime('%m')
        date_col_index = df.columns.get_loc("Invoice Date")
        df.insert(date_col_index + 1, "Month", month_series)

# 合并去重
df_combined = pd.concat([df_main, df_append], ignore_index=True)
df_combined = df_combined[df_combined["Invoice Number"].notna()]
df_combined["Invoice Number"] = df_combined["Invoice Number"].astype(str).str.strip()
df_combined = df_combined.drop_duplicates(subset="Invoice Number")

# 格式化日期列为 YYYY-MM-DD
df_combined["Invoice Date"] = pd.to_datetime(df_combined["Invoice Date"], errors='coerce')
df_combined["Invoice Date"] = df_combined["Invoice Date"].dt.strftime('%Y-%m-%d')

# 排序
df_combined.sort_values(by=["Month", "Invoice Date"], inplace=True)

# 创建输出目录（如不存在）
os.makedirs(output_dir, exist_ok=True)

# 保存为 Excel
df_combined.to_excel(output_file, index=False)

# 格式化 'Invoice Amount' 为美元
wb = load_workbook(output_file)
ws = wb.active

invoice_col_letter = None
for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == "Invoice Amount":
            invoice_col_letter = cell.column_letter
            break

if invoice_col_letter:
    for row in ws.iter_rows(min_row=2, min_col=ws[invoice_col_letter + "1"].column, max_col=ws[invoice_col_letter + "1"].column):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    wb.save(output_file)
    print(f"✅ 合并完成，已保存为：{output_file}")
else:
    print("⚠️ 未找到 'Invoice Amount' 列，未设置货币格式。")
